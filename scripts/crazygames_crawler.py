#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
CrazyGames爬虫脚本
用于爬取CrazyGames网站的游戏数据并生成Excel文件
"""

import os
import re
import json
import time
import random
import requests
import pandas as pd
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import pytz

# 配置
class Config:
    # 基础URL
    BASE_URL = 'https://www.crazygames.com'
    
    # 游戏列表页URL
    GAMES_LIST_URL = 'https://www.crazygames.com/t/html5'
    
    # 分类页URL模板
    CATEGORY_URL_TEMPLATE = 'https://www.crazygames.com/c/{}'
    
    # 请求头
    HEADERS = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Cache-Control': 'max-age=0',
    }
    
    # 请求间隔(秒)，避免过于频繁的请求
    REQUEST_DELAY_MIN = 1
    REQUEST_DELAY_MAX = 3
    
    # 重试次数
    MAX_RETRIES = 3
    
    # 重试延迟(秒)
    RETRY_DELAY = 2
    
    # 输出Excel文件名
    OUTPUT_EXCEL = 'crazygames_games.xlsx'
    
    # 是否启用调试模式
    DEBUG = True


class CrazyGamesCrawler:
    def __init__(self):
        """初始化爬虫"""
        self.config = Config()
        self.session = requests.Session()
        self.session.headers.update(self.config.HEADERS)
        
        # 创建输出目录
        os.makedirs('output', exist_ok=True)
        
        # 创建脚本目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        self.output_dir = os.path.join(script_dir, 'output')
        self.import_file = os.path.join(script_dir, 'games_for_import.json')
        os.makedirs(self.output_dir, exist_ok=True)
        
        self.log("CrazyGames爬虫初始化完成")
    
    def log(self, message, level='INFO'):
        """日志记录"""
        if level == 'DEBUG' and not self.config.DEBUG:
            return
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{timestamp}] [{level}] {message}")
    
    def fetch_url(self, url, retries=0):
        """获取URL内容"""
        try:
            self.log(f"正在获取: {url}", 'DEBUG')
            
            # 添加随机延迟
            delay = random.uniform(self.config.REQUEST_DELAY_MIN, self.config.REQUEST_DELAY_MAX)
            time.sleep(delay)
            
            response = self.session.get(url)
            response.raise_for_status()
            
            return response.text
        except requests.RequestException as e:
            self.log(f"获取URL失败: {url}, 错误: {str(e)}", 'ERROR')
            
            # 如果未超过最大重试次数，则重试
            if retries < self.config.MAX_RETRIES:
                self.log(f"正在重试({retries + 1}/{self.config.MAX_RETRIES}): {url}", 'WARNING')
                time.sleep(self.config.RETRY_DELAY)
                return self.fetch_url(url, retries + 1)
            
            # 超过最大重试次数，返回空字符串
            return ""
    
    def get_games_by_category(self, category='', limit=10):
        """获取指定分类的游戏列表"""
        # 构建URL
        url = self.config.GAMES_LIST_URL
        
        if category:
            # 检查是否是完整URL
            if category.startswith('http'):
                url = category
            # 检查是否是分类名称
            elif not category.isdigit():
                url = self.config.CATEGORY_URL_TEMPLATE.format(category)
            # 如果是数字，可能是分类索引，但我们现在知道分类是字符串，所以直接使用HTML5页面
            else:
                self.log(f"分类'{category}'似乎是数字，但CrazyGames使用字符串分类标识符。使用默认HTML5页面。", 'WARNING')
        
        self.log(f"正在从 {url} 获取游戏列表...")
        
        # 获取分类页HTML
        html = self.fetch_url(url)
        if not html:
            self.log(f"获取游戏列表失败: 无法获取页面内容", 'ERROR')
            return []
        
        # 保存HTML以便分析
        with open(os.path.join(self.output_dir, 'crazygames_html.txt'), 'w', encoding='utf-8') as f:
            f.write(html)
        self.log("已保存HTML到crazygames_html.txt文件，用于分析网站结构")
        
        soup = BeautifulSoup(html, 'html.parser')
        
        # 尝试多种可能的选择器
        possible_selectors = [
            '.game-list-item',   # 新的选择器
            '.game-grid-item',   # 原始选择器
            '.game-item',        # 常见的游戏项命名
            '.game-card',        # 常见的卡片命名
            '.game',             # 简单的游戏类
            '.card',             # 通用卡片类
            'article',           # 可能的文章元素
            '.thumbnail',        # 缩略图容器
            'a[href*="/game/"]'  # 包含/game/的链接
        ]
        
        game_cards = []
        for selector in possible_selectors:
            elements = soup.select(selector)
            if elements:
                self.log(f"使用选择器 '{selector}' 找到 {len(elements)} 个元素")
                game_cards = elements
                break
        
        if not game_cards:
            self.log("未能找到游戏卡片元素，尝试查找所有游戏链接", 'WARNING')
            # 尝试查找所有游戏链接
            game_links = soup.select('a[href*="/game/"]')
            if game_links:
                self.log(f"找到 {len(game_links)} 个游戏链接")
                game_cards = game_links
            else:
                self.log("未能找到游戏元素，请检查网站结构", 'ERROR')
                return []
        
        # 限制游戏数量
        limited_cards = game_cards[:limit]
        
        self.log(f"找到 {len(limited_cards)} 个游戏卡片")
        
        # 存储游戏数据
        games = []
        
        # 处理每个游戏卡片
        for index, card in enumerate(limited_cards):
            try:
                # 打印当前卡片的HTML结构，帮助调试
                self.log(f"处理游戏卡片 {index+1}:", 'DEBUG')
                self.log(str(card)[:200] + "...", 'DEBUG')  # 只打印前200个字符
                
                # 提取基本信息
                href = card.get('href', '')
                
                # 如果是链接元素，直接获取文本作为标题
                title_element = card.select_one('.game-list-item-title, .game-title, h3, h2, .title') or card
                title = title_element.get_text(strip=True)
                
                # 提取缩略图URL
                thumbnail_url = ''
                img_element = card.select_one('img')
                if img_element:
                    # 尝试多个可能的属性
                    for attr in ['src', 'data-src', 'data-lazy-src', 'data-original']:
                        if img_element.get(attr):
                            thumbnail_url = img_element.get(attr)
                            break
                    
                    # 如果找到了缩略图，确保它是完整URL
                    if thumbnail_url and not thumbnail_url.startswith(('http://', 'https://')):
                        thumbnail_url = urljoin(self.config.BASE_URL, thumbnail_url)
                    
                    self.log(f"找到缩略图: {thumbnail_url}", 'DEBUG')
                
                # 构建游戏详情页URL
                game_url = urljoin(self.config.BASE_URL, href) if href and not href.startswith('http') else href
                
                # 如果没有有效的URL，跳过此游戏
                if not game_url:
                    self.log(f"跳过游戏 {title}: 无法获取游戏URL", 'WARNING')
                    continue
                
                self.log(f"正在处理游戏 ({index + 1}/{len(limited_cards)}): {title}")
                
                # 获取游戏详情
                game_details = self.get_game_details(game_url)
                
                # 合并数据
                game_data = {
                    'title': title,
                    'url': game_url,
                    'thumbnailUrl': thumbnail_url,
                    **game_details
                }
                
                games.append(game_data)
                self.log(f"成功添加游戏: {title}", 'INFO')
                
            except Exception as e:
                self.log(f"处理游戏卡片时出错: {str(e)}", 'ERROR')
                import traceback
                self.log(traceback.format_exc(), 'ERROR')
        
        return games
    
    def get_game_details(self, url):
        """获取游戏详情"""
        self.log(f"正在获取游戏详情: {url}", 'DEBUG')
        
        # 获取游戏详情页HTML
        html = self.fetch_url(url)
        if not html:
            self.log(f"获取游戏详情失败: 无法获取页面内容", 'ERROR')
            return self._get_empty_game_details(url)
        
        soup = BeautifulSoup(html, 'html.parser')
        
        # 提取游戏详情
        description_selectors = ['.game-description', '.description', 'meta[name="description"]', 'meta[property="og:description"]']
        instructions_selectors = ['.game-instructions', '.instructions', '.how-to-play']
        category_selectors = ['.game-info-category a', '.category a', '.game-category', '.game-tags a']
        
        # 尝试多个选择器获取描述
        description = ''
        for selector in description_selectors:
            element = soup.select_one(selector)
            if element:
                if element.name == 'meta':
                    description = element.get('content', '')
                else:
                    description = element.get_text(strip=True)
                if description:
                    break
        
        # 尝试多个选择器获取说明
        instructions = ''
        for selector in instructions_selectors:
            element = soup.select_one(selector)
            if element:
                instructions = element.get_text(strip=True)
                if instructions:
                    break
        
        # 尝试多个选择器获取分类
        category = ''
        for selector in category_selectors:
            elements = soup.select(selector)  # 获取所有匹配的元素
            for element in elements:
                category_text = element.get_text(strip=True)
                if category_text and category_text.lower() in self.get_predefined_categories():
                    category = category_text
                    break
            if category:
                break
        
        # 如果没有找到有效分类，使用默认分类
        if not category:
            # 根据游戏标题或描述推断分类
            category = self.guess_category(url, description)
        
        # 确保分类首字母大写
        category = category.capitalize()
        
        # 提取评分
        rating_selectors = ['.game-rating-value', '.rating-value', '.rating']
        rating = 0
        for selector in rating_selectors:
            element = soup.select_one(selector)
            if element:
                try:
                    rating_text = element.get_text(strip=True)
                    # 处理可能的百分比格式
                    if '%' in rating_text:
                        rating = float(rating_text.replace('%', '')) / 20  # 转换为5分制
                    else:
                        rating = float(rating_text)
                    break
                except (ValueError, TypeError):
                    continue
        
        # 如果没有评分，设置默认评分
        if not rating:
            rating = 3.5
        
        # 确保评分在1-5之间
        rating = max(1.0, min(5.0, rating))
        
        # 提取嵌入URL
        embed_url = self._get_embed_url(soup, url)
        
        # 提取游戏ID
        game_id = url.split('/')[-1]
        
        # 尝试从脚本中提取更多游戏数据
        game_data = self._extract_game_data_from_scripts(soup)
        
        # 如果没有找到缩略图，尝试从og:image元标签获取
        thumbnail_url = game_data.get('thumbnailUrl', '')
        if not thumbnail_url:
            og_image = soup.select_one('meta[property="og:image"]')
            if og_image:
                thumbnail_url = og_image.get('content', '')
        
        return {
            'id': game_id,
            'description': description or f"Play {url.split('/')[-1].replace('-', ' ').title()} online for free!",
            'instructions': instructions or "Use your mouse and keyboard to play the game.",
            'category': category,
            'rating': rating,
            'embedUrl': embed_url,
            'thumbnailUrl': thumbnail_url,
            **game_data
        }
    
    def get_predefined_categories(self):
        """获取预定义的游戏分类"""
        return {
            'action', 'adventure', 'arcade', 'puzzle', 'racing', 
            'sports', 'strategy', 'shooting', 'horror', 'simulation'
        }
    
    def guess_category(self, url, description):
        """根据URL和描述推断游戏分类"""
        text = f"{url} {description}".lower()
        
        # 分类关键词映射
        category_keywords = {
            'action': ['action', 'fight', 'battle', 'combat', 'shoot'],
            'adventure': ['adventure', 'explore', 'quest', 'journey'],
            'arcade': ['arcade', 'classic', 'retro'],
            'puzzle': ['puzzle', 'match', 'solve', 'brain'],
            'racing': ['racing', 'drive', 'car', 'speed', 'drift'],
            'sports': ['sports', 'football', 'basketball', 'soccer'],
            'strategy': ['strategy', 'build', 'manage', 'tower'],
            'shooting': ['shooting', 'shooter', 'gun', 'fps'],
            'horror': ['horror', 'scary', 'survival', 'zombie'],
            'simulation': ['simulation', 'simulator', 'realistic']
        }
        
        # 计算每个分类的匹配度
        matches = {}
        for category, keywords in category_keywords.items():
            matches[category] = sum(1 for keyword in keywords if keyword in text)
        
        # 返回匹配度最高的分类，如果没有匹配则返回 'Action'
        best_category = max(matches.items(), key=lambda x: x[1])[0] if any(matches.values()) else 'Action'
        return best_category.capitalize()
    
    def _extract_game_data_from_scripts(self, soup):
        """从页面脚本中提取游戏数据"""
        game_data = {}
        
        # 查找包含游戏数据的脚本
        scripts = soup.select('script')
        
        for script in scripts:
            script_text = script.string
            if not script_text:
                continue
            
            # 查找window.game或window.crazygames.game对象
            game_data_match = re.search(r'window\.game\s*=\s*({.*?});', script_text, re.DOTALL)
            if not game_data_match:
                game_data_match = re.search(r'window\.crazygames\.game\s*=\s*({.*?});', script_text, re.DOTALL)
            
            if game_data_match:
                try:
                    # 尝试解析JSON
                    data_str = game_data_match.group(1)
                    # 替换单引号为双引号
                    data_str = re.sub(r'\'', '"', data_str)
                    # 处理JavaScript对象中的函数和非标准JSON
                    data_str = re.sub(r'(\w+):', r'"\1":', data_str)
                    data_str = re.sub(r',\s*}', '}', data_str)
                    
                    # 尝试解析，但忽略错误
                    try:
                        parsed_data = json.loads(data_str)
                        game_data.update(parsed_data)
                    except json.JSONDecodeError:
                        pass
                except Exception as e:
                    self.log(f"解析游戏数据时出错: {str(e)}", 'ERROR')
        
        return game_data
    
    def _get_embed_url(self, soup, fallback_url):
        """获取游戏嵌入URL"""
        # 尝试从iframe中提取
        iframe = soup.select_one('iframe.game-iframe, iframe[id*="game"], iframe[src*="game"], iframe[src*="embed"]')
        if iframe:
            src = iframe.get('src')
            if src:
                return src
        
        # 尝试构造嵌入URL
        game_slug = fallback_url.split('/')[-1]
        if game_slug and 'crazygames.com' in fallback_url:
            embed_url = f"https://www.crazygames.com/embed/{game_slug}"
            return embed_url
        
        # 如果都失败了，返回原始URL
        return fallback_url
    
    def _get_empty_game_details(self, url):
        """返回空的游戏详情"""
        return {
            'id': url.split('/')[-1],
            'description': '',
            'instructions': '',
            'category': '',
            'rating': 0,
            'embedUrl': url
        }
    
    def _convert_to_site_format(self, game):
        """将爬取的游戏数据转换为网站使用的格式"""
        now = datetime.now(pytz.timezone('Asia/Shanghai'))
        
        # 从URL中提取ID
        game_id = game.get('id', '')
        if not game_id:
            url = game.get('url', '')
            game_id = url.split('/')[-1] if url else ''
        
        # 基本游戏信息
        site_game = {
            'title': game.get('title', ''),
            'url': game.get('url', ''),
            'thumbnailUrl': game.get('thumbnailUrl', ''),
            'id': game_id,
            'description': game.get('description', ''),
            'instructions': game.get('instructions', game.get('description', '')),
            'category': game.get('category', 'Action'),
            'rating': float(game.get('rating', 3.5)),
            'embedUrl': game.get('embedUrl', '')
        }
        
        # 根据游戏ID的哈希值决定游戏状态
        # 使用最后一个字符的ASCII值，保证分配均匀
        last_char = game_id[-1] if game_id else 'a'
        ascii_value = ord(last_char.lower())
        
        # 根据ASCII值对3取余，平均分配三种状态：
        # 0: 新游戏 (isNew = true, isFeatured = false)
        # 1: 推荐游戏 (isNew = false, isFeatured = true)
        # 2: 普通游戏 (isNew = false, isFeatured = false)
        state = ascii_value % 3
        
        site_game['isNew'] = state == 0
        site_game['isFeatured'] = state == 1
        
        # 添加状态和时间戳
        if site_game['isNew']:
            site_game['status'] = 'New'
        elif site_game['isFeatured']:
            site_game['status'] = 'Featured'
        else:
            site_game['status'] = 'Normal'
            
        site_game['addedDate'] = now.isoformat()
        site_game['lastUpdated'] = now.isoformat()
        
        return site_game
    
    def export_to_json(self, games):
        """导出游戏数据到JSON文件，适合网站导入"""
        if not games:
            self.log("没有游戏数据可导出", 'WARNING')
            return False
        
        try:
            # 转换为网站格式
            site_format_games = []
            for game in games:
                site_game = self._convert_to_site_format(game)
                site_format_games.append(site_game)
            
            # 保存JSON文件到output目录
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = os.path.join(self.output_dir, f"crazygames_data_{timestamp}.json")
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(site_format_games, f, ensure_ascii=False, indent=2)
            
            self.log(f"数据已成功导出到 {filename}", 'INFO')
            
            # 创建admin/crawler-import.html可以使用的导入文件
            with open(self.import_file, 'w', encoding='utf-8') as f:
                json.dump(site_format_games, f, ensure_ascii=False, indent=2)
            
            self.log(f"导入文件已创建: {self.import_file}", 'INFO')
            return True
        except Exception as e:
            self.log(f"导出JSON失败: {str(e)}", 'ERROR')
            import traceback
            self.log(traceback.format_exc(), 'ERROR')
            return False
    
    def export_to_excel(self, games):
        """将游戏数据导出到Excel文件"""
        if not games:
            self.log("没有游戏数据可导出", 'WARNING')
            return False
        
        try:
            # 创建一个新的Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "游戏数据"
            
            # 添加表头
            headers = ['游戏名称', '游戏描述', '游戏分类', '嵌入URL']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
            
            # 添加数据
            for row_num, game in enumerate(games, 2):
                ws.cell(row=row_num, column=1).value = game.get('title', '')
                ws.cell(row=row_num, column=2).value = game.get('description', '')
                ws.cell(row=row_num, column=3).value = game.get('category', '')
                ws.cell(row=row_num, column=4).value = game.get('embedUrl', '')
            
            # 保存Excel文件
            filename = f"crazygames_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            self.log(f"数据已成功导出到 {filename}", 'INFO')
            return True
        
        except Exception as e:
            self.log(f"导出Excel失败: {str(e)}", 'ERROR')
            return False


def load_existing_games():
    """加载现有的游戏数据"""
    try:
        with open('scripts/games_for_import.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data.get('games', [])
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def merge_games(existing_games, new_games):
    """合并新旧游戏数据，使用id作为唯一标识"""
    games_dict = {game['id']: game for game in existing_games}
    
    # 更新现有游戏或添加新游戏
    for game in new_games:
        games_dict[game['id']] = game
    
    return list(games_dict.values())

def save_games_data(games):
    """保存游戏数据，包括更新ID和时间戳"""
    now = datetime.now(pytz.timezone('Asia/Shanghai'))
    data = {
        "updateId": now.strftime("%Y%m%d_%H%M%S"),
        "lastUpdated": now.isoformat(),
        "games": games
    }
    
    with open('scripts/games_for_import.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def main():
    # 加载现有游戏数据
    existing_games = load_existing_games()
    
    # 初始化爬虫
    crawler = CrazyGamesCrawler()
    
    # 显示可用分类
    predefined_categories = [
        {'name': '休闲游戏', 'slug': 'casual'},
        {'name': '动作游戏', 'slug': 'action'},
        {'name': '冒险游戏', 'slug': 'adventure'},
        {'name': '益智游戏', 'slug': 'puzzle'},
        {'name': '射击游戏', 'slug': 'shooting'},
        {'name': '体育游戏', 'slug': 'sports'},
        {'name': '多人游戏', 'slug': 'multiplayer'},
        {'name': 'HTML5游戏', 'slug': 'html5'}
    ]
    
    print("\n可用分类:")
    for index, category in enumerate(predefined_categories, 1):
        print(f"{index}. {category['name']} ({category['slug']})")
    
    # 获取用户输入
    print("\n请选择爬取选项:")
    category_input = input("分类 (输入编号或分类名，留空表示所有分类): ")
    limit_input = input("爬取数量 (默认10): ")
    
    # 处理分类输入
    selected_category = ''
    if category_input:
        try:
            index = int(category_input) - 1
            if 0 <= index < len(predefined_categories):
                selected_category = predefined_categories[index]['slug']
            else:
                selected_category = category_input
        except ValueError:
            # 作为分类名处理
            selected_category = category_input
    
    # 处理数量输入
    try:
        limit = int(limit_input) if limit_input else 10
    except ValueError:
        limit = 10
    
    print(f"\n开始爬取: 分类={selected_category or '所有'}, 数量={limit}")
    
    # 爬取游戏
    games = crawler.get_games_by_category(selected_category, limit)
    
    # 合并游戏数据
    all_games = merge_games(existing_games, games)
    
    # 保存更新后的数据
    save_games_data(all_games)
    
    # 导出数据
    if games:
        # 导出到Excel
        crawler.export_to_excel(games)
        
        # 导出到JSON，适合网站导入
        crawler.export_to_json(games)
        
        print(f"\n成功爬取 {len(games)} 个游戏并导出")
    else:
        print("\n未爬取到任何游戏数据")
    
    print("\n爬取完成!")


if __name__ == "__main__":
    main()
